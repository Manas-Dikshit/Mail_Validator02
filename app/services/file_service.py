from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config.settings import get_settings
from app.models.schemas import EmailValidationResult

logger = logging.getLogger(__name__)
settings = get_settings()


class NoEmailColumnFoundError(Exception):
    pass


class MultipleEmailColumnsError(Exception):
    """Raised when auto-detection finds more than one plausible email
    column and the caller hasn't specified which one to use."""

    def __init__(self, candidates: list[str]):
        self.candidates = candidates
        super().__init__(f"Multiple possible email columns found: {candidates}")


def load_dataframe(file_path: str | Path) -> pd.DataFrame:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(path, dtype=str, keep_default_na=False, na_values=[""])
    if suffix in (".xlsx", ".xlsm"):
        return pd.read_excel(path, dtype=str, engine="openpyxl")
    raise ValueError(f"Unsupported file type: {suffix}. Use .csv or .xlsx")


def load_dataframe_chunks(file_path: str | Path, chunk_size: Optional[int] = None):
    """Generator of DataFrame chunks for memory-efficient processing of
    very large CSV files. XLSX is loaded whole (openpyxl doesn't support
    true streaming reads via pandas), but that's fine since XLSX row
    counts are typically much smaller than raw CSV exports."""
    chunk_size = chunk_size or settings.csv_chunk_size
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from pd.read_csv(
            path, dtype=str, keep_default_na=False, na_values=[""], chunksize=chunk_size
        )
    else:
        df = load_dataframe(path)
        for start in range(0, len(df), chunk_size):
            yield df.iloc[start : start + chunk_size]


def detect_email_column(
    columns: list[str], preferred: Optional[str] = None
) -> str:
    """Detect the email column by matching against known candidate names
    (case-insensitive, punctuation-tolerant). Raises if none or multiple
    plausible columns are found, so the caller can prompt the user."""
    if preferred:
        if preferred in columns:
            return preferred
        raise NoEmailColumnFoundError(f"Requested column '{preferred}' not found")

    def normalize(name: str) -> str:
        return "".join(ch for ch in name.lower() if ch.isalnum())

    candidate_norms = {normalize(c) for c in settings.email_column_candidates}

    matches = [c for c in columns if normalize(c) in candidate_norms]

    if not matches:
        # Fallback: any column whose normalized name *contains* "email" or "mail"
        matches = [
            c for c in columns if "email" in normalize(c) or normalize(c).endswith("mail")
        ]

    if not matches:
        raise NoEmailColumnFoundError(
            f"No email column detected among: {columns}. "
            "Supported names include: Email, Email Address, email_id, e-mail, etc."
        )
    if len(matches) > 1:
        raise MultipleEmailColumnsError(matches)

    return matches[0]


def build_output_dataframe(
    original_df: pd.DataFrame,
    email_column: str,
    results: list[EmailValidationResult],
) -> pd.DataFrame:
    """Append validation result columns onto the original dataframe,
    preserving every original column exactly as-is."""
    result_records = []
    for r in results:
        flat = r.to_flat_dict()
        result_records.append(
            {
                "Normalized Email": flat["normalized_email"],
                "Validation Status": flat["validation_status"],
                "Primary Tag": flat["primary_tag"],
                "Secondary Tag": flat["secondary_tag"],
                "Syntax Valid": flat["syntax_valid"],
                "Domain Exists": flat["domain_exists"],
                "MX Exists": flat["mx_exists"],
                "DNS Status": flat["dns_status"],
                "Disposable": flat["disposable"],
                "Role Based": flat["role_based"],
                "Free Provider": flat["free_provider"],
                "Free Provider Name": flat["free_provider_name"],
                "Business Email": flat["business_email"],
                "International Email": flat["international"],
                "Typo Suggestion": flat["typo_suggestion"],
                "Catch All Possible": flat["catch_all_status"],
                "Mailbox Status": flat["mailbox_status"],
                "Spam Score": flat["spam_score"],
                "Risk Score": flat["risk_score"],
                "Reason": flat["reason"],
                "Recommendation": flat["recommendation"],
                "Deliverability Score": flat["deliverability_score"],
                "Send Decision": flat["send_decision"],
                "Domain Reputation Flags": ", ".join(flat["domain_reputation_flags"]),
                "Has SPF": flat["has_spf"],
                "Has DKIM Indicator": flat["has_dkim_indicator"],
                "Has DMARC": flat["has_dmarc"],
                "MX Provider": flat["mx_provider"],
                "DNS Response Time (ms)": flat["dns_response_time_ms"],
                "Validation Timestamp": flat["validation_timestamp"],
                "Duplicate Count": flat["duplicate_count"],
            }
        )

    results_df = pd.DataFrame(result_records)
    output_df = pd.concat(
        [original_df.reset_index(drop=True), results_df.reset_index(drop=True)], axis=1
    )
    return output_df


_STATUS_COLORS = {
    "VALID": "C6EFCE",
    "INVALID": "FFC7CE",
    "SUSPICIOUS": "FFEB9C",
    "RISKY": "FFEB9C",
    "UNKNOWN": "D9D9D9",
}


def write_output_xlsx(df: pd.DataFrame, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Validation Results")
        worksheet = writer.sheets["Validation Results"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        # Auto-size columns (bounded, to avoid pathological widths)
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max(
                [len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str).head(500)]
            )
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
                max(12, max_len + 2), 45
            )

        worksheet.freeze_panes = "A2"

        if "Validation Status" in df.columns:
            status_col_idx = list(df.columns).index("Validation Status") + 1
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=status_col_idx)
                value = str(cell.value or "")
                for prefix, color in _STATUS_COLORS.items():
                    if value.startswith(prefix):
                        cell.fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid"
                        )
                        break

    logger.info("Wrote %d rows to %s", len(df), output_path)
    return output_path
